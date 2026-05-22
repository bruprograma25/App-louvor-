#!/usr/bin/env python
import os
import sys
import json
from openpyxl import load_workbook
import requests

BASE_API = os.environ.get('API_URL', 'http://127.0.0.1:5000/api')
DATA_DIR = os.path.join(os.getcwd(), 'data')

def read_sheet(path):
    wb = load_workbook(path, data_only=True)
    sheet = wb[wb.sheetnames[0]]
    rows = list(sheet.rows)
    if not rows:
        return []
    headers = [str(cell.value).strip() if cell.value is not None else '' for cell in rows[0]]
    data = []
    for r in rows[1:]:
        item = {}
        for h, cell in zip(headers, r):
            key = h.strip().lower() if h else ''
            item[key] = cell.value
        data.append(item)
    return data


def import_members(rows):
    created = 0
    for idx, row in enumerate(rows):
        # try to extract common fields
        full_name = row.get('nome') or row.get('full name') or row.get('nome completo') or row.get('name') or ''
        email = row.get('email') or row.get('e-mail') or ''
        role = row.get('cargo') or row.get('role') or row.get('função') or ''
        image = row.get('imagem') or row.get('image') or ''
        password = row.get('password') or row.get('senha') or 'changeme123'
        if not email:
            print(f"Skipping row {idx+1}: missing email")
            continue
        payload = {
            'full_name': str(full_name) if full_name is not None else '',
            'email': str(email),
            'password': str(password),
            'role': str(role) if role is not None else 'member',
            'status': str(row.get('status') or row.get('situação') or row.get('situacao') or 'Apto'),
            'image': str(image) if image else None,
        }
        try:
            res = requests.post(f"{BASE_API}/users", json=payload, timeout=10)
            if res.status_code in (200,201):
                created += 1
                print(f"Created user: {email}")
            else:
                print(f"Failed to create {email}: {res.status_code} {res.text}")
        except Exception as e:
            print("Error creating user", e)
    print(f"Imported members: {created}")


def import_processes(rows):
    created = 0
    for idx, row in enumerate(rows):
        title = row.get('titulo') or row.get('title') or row.get('assunto') or f"Processo {idx+1}"
        description = row.get('descricao') or row.get('description') or row.get('descrição') or ''
        category = row.get('categoria') or row.get('category') or ''
        due = row.get('prazo') or row.get('due') or row.get('data') or ''
        link = row.get('link') or row.get('recurso') or ''
        payload = {
            'title': str(title) if title is not None else '',
            'date': str(due) if due is not None else '',
            'notes': str(description) if description is not None else '',
        }
        try:
            res = requests.post(f"{BASE_API}/ministrations", json=payload, timeout=10)
            if res.status_code in (200,201):
                created += 1
                print(f"Created process/ministration: {title}")
            else:
                print(f"Failed to create process {title}: {res.status_code} {res.text}")
        except Exception as e:
            print("Error creating process", e)
    print(f"Imported processes/ministrations: {created}")


def main():
    files = [f for f in os.listdir(DATA_DIR) if f.lower().endswith(('.xlsx', '.xlsm', '.csv'))]
    if not files:
        print('No data files found in data/. Put your .xlsx/.xlsm/.csv files there and retry.')
        return
    for file in files:
        path = os.path.join(DATA_DIR, file)
        print('\nProcessing', file)
        try:
            rows = read_sheet(path) if not file.lower().endswith('.csv') else []
            # fallback for CSV
            if file.lower().endswith('.csv'):
                import csv
                with open(path, encoding='utf-8') as fh:
                    reader = csv.DictReader(fh)
                    rows = [dict((k.strip().lower(), v) for k,v in row.items()) for row in reader]
        except Exception as e:
            print('Failed to read file', file, e)
            continue

        # Heuristic: if there's an email column -> members
        keys = set().union(*(r.keys() for r in rows)) if rows else set()
        if any('email' in k for k in keys):
            print('Detected members file, importing users...')
            import_members(rows)
        else:
            print('Assuming processes/ministrations file, importing as ministrations...')
            import_processes(rows)

if __name__ == '__main__':
    main()
